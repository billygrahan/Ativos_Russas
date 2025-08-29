import heapq
import math
import os
import random
import time
import sys
import copy
# pip install psutil openpyxl
import psutil
import openpyxl

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from Algoritmos.BFS import bfs
from Algoritmos.DFS import dfs
from Algoritmos.BCU import bcu
from Algoritmos.A_ESTRELA import a_estrela, heuristica_euclidiana, heuristica_haversiana

qtd_ativos = 0
nome_arquivo_entrada = ''
nome_arquivo_saida = ''

melhor_rota = []
distancia_entre_ativos = {}
caminho_entre_ativos = []

graph_dist = {}
graph_Coordenadas = {}
qtd_vertices = 0

inicio = 5
ativos = []

alg = '' 
iteracoes = 1000

def algoritmo(origem, destino):
    if alg == 'A_Estrela_Haversiano':
        return a_estrela(graph_dist, graph_Coordenadas, origem, destino, heuristica_haversiana)
    elif alg == 'A_Estrela_Euclidiano':
        return a_estrela(graph_dist, graph_Coordenadas, origem, destino, heuristica_euclidiana)
    elif alg == 'BFS':
        return bfs(graph_dist, origem, destino)
    elif alg == 'DFS':
        return dfs(graph_dist, origem, destino)
    elif alg == 'BCU':
        return bcu(graph_dist, origem, destino)
 
def Carrega_Dados():
    global graph_dist, graph_Coordenadas, qtd_vertices, ativos, inicio, distancia_entre_ativos

    filepath = os.path.join(os.path.dirname(__file__), f"{nome_arquivo_entrada}.gr")
    with open(filepath, "r") as arquivo:
        for linha in arquivo:
            linha_dividida = linha.split()
                
            vertice_origem = int(linha_dividida[1])
            vertice_destino = int(linha_dividida[2])
            vertice_distancia = int(linha_dividida[3])
            
            if vertice_origem not in graph_dist.keys():
                graph_dist[vertice_origem] = {vertice_destino:vertice_distancia}
            else:
                graph_dist[vertice_origem][vertice_destino] = vertice_distancia

    filepath_co = os.path.join(os.path.dirname(__file__), f"{nome_arquivo_entrada}.co")
    with open(filepath_co, "r") as arquivo:
        for linha in arquivo:
            linha_dividida = linha.split()
            
            vertice_origem = int(linha_dividida[1])
            coordenada1 = float(linha_dividida[2])
            coordenada2 = float(linha_dividida[3])
            
            graph_Coordenadas[vertice_origem] = (coordenada1, coordenada2)

def Roleta(roleta):
    objetos = [obj for obj, dist in roleta]
    distancias = [dist for obj, dist in roleta]

    for obj, dist in roleta:
        if dist == 0:
            return obj

    pesos = [1/(d if d > 0 else 1e-6) for d in distancias]
    total_peso = sum(pesos)
    probabilidades = [p/total_peso for p in pesos]

    escolhido = random.choices(objetos, weights=probabilidades, k=1)[0]
    return escolhido

def recalcula_distancias(melhor_rota_copia):
    global distancia_entre_ativos , inicio, ativos , graph_dist, graph_Coordenadas
    
    for i in range(len(melhor_rota_copia)-1):
        origem = melhor_rota_copia[i][0]
        destino = melhor_rota_copia[i+1][0]
        
        if destino not in distancia_entre_ativos[origem]:
            heuristica = algoritmo(origem, destino)
            rota_encontrada, distancia_percorrida, quantidade_nos_expandidos, fator_ramificacao = heuristica
            distancia_entre_ativos[origem][destino] = distancia_percorrida
            caminho_entre_ativos.append(rota_encontrada)
            melhor_rota_copia[i+1] = (destino, melhor_rota_copia[i][1] + distancia_percorrida)
        else:
            distancia_percorrida = distancia_entre_ativos[origem][destino]
            melhor_rota_copia[i+1] = (destino, melhor_rota_copia[i][1] + distancia_percorrida)

def Gerar_Solucao():
    global graph_dist, graph_Coordenadas, qtd_vertices, ativos, inicio, distancia_entre_ativos, melhor_rota, caminho_entre_ativos, alg, iteracoes

    melhor_rota_copia = []
    ativos_copia = ativos.copy()
    origem = inicio

    melhor_rota_copia.append((origem, 0))

    while len(ativos_copia) > 0:
        roleta = []

        for ativo in ativos_copia:
            if ativo not in distancia_entre_ativos[origem]:
                rota_encontrada, distancia, quantidade_nos_expandidos, fator_ramificacao = algoritmo(origem, ativo)
                caminho_entre_ativos.append(rota_encontrada)
                distancia_entre_ativos[origem][ativo] = distancia
                roleta.append((ativo, distancia))
            else:
                roleta.append((ativo, distancia_entre_ativos[origem][ativo]))

        elemento = Roleta(roleta)

        melhor_rota_copia.append((elemento, 0))

        ativos_copia.remove(elemento)
        origem = elemento

    # Adiciona o ponto de origem ao final da rota
    if inicio not in distancia_entre_ativos[origem]:
        rota_encontrada, distancia, quantidade_nos_expandidos, fator_ramificacao = algoritmo(origem, inicio)
        caminho_entre_ativos.append(rota_encontrada)
        distancia_entre_ativos[origem][inicio] = distancia
        melhor_rota_copia.append((inicio, 0))
    else:
        melhor_rota_copia.append((inicio, 0))

    recalcula_distancias(melhor_rota_copia)
    return melhor_rota_copia

def melhorar_Rota():
    global graph_dist, graph_Coordenadas, qtd_vertices, ativos, inicio, distancia_entre_ativos, melhor_rota, caminho_entre_ativos, alg, iteracoes
    
    for i in range(iteracoes):
        melhor_rota_copia = Gerar_Solucao()

        if melhor_rota == [] or melhor_rota_copia[-1][1] < melhor_rota[-1][1]:
            melhor_rota = melhor_rota_copia
    

def carrega_teste():
    global graph_dist, graph_Coordenadas, qtd_vertices, ativos, inicio, distancia_entre_ativos, melhor_rota, caminho_entre_ativos, alg, iteracoes
    
    #algoritmos = ['BFS', 'DFS', 'BCU', 'A_Estrela_Euclidiano', 'A_Estrela_Haversiano']
    planilha = openpyxl.Workbook()
    del planilha['Sheet']
    planilha.create_sheet(alg)

    pagina = planilha[alg]
    pagina.sheet_format.baseColWidth = 30
    pagina.append(['Índice', 'inicio', 'Distancia', 'ativos', 'melhor rota', 'caminho', 'Tempo'])

    sementes = [286208,998628,34825,342117,67982,148086,513282,306647,93844,609550,582231,725359,408128,481117,450284,102701,938521,105559]
    #print("     INICIO      ;      ATIVOS      ;      MELHOR ROTA ENTRE ATIVOS      ;      CAMINHO TOTAL     ")
    for c in range(3): # qtd de testes
        random.seed(sementes[c])
        qtd_vertices = len(graph_Coordenadas)

        inicio = random.randint(1, qtd_vertices)
        ativos = random.sample(range(1, qtd_vertices + 1), qtd_ativos)

        melhor_rota = []
        caminho_entre_ativos = []
        distancia_entre_ativos = {}

        distancia_entre_ativos[inicio] = {}
        for ativo in ativos:
            distancia_entre_ativos[ativo] = {}

        inicio_tempo = time.time()
        melhorar_Rota()
        tempo_Rota = time.time() - inicio_tempo

        caminho_completo = [str(melhor_rota[0][0])]
        for i in range(len(melhor_rota)-1):
            # Procura o caminho entre melhor_rota[i][0] e melhor_rota[i+1][0]
            for caminho in caminho_entre_ativos:
                if melhor_rota[i][0] == caminho[0] and melhor_rota[i+1][0] == caminho[-1]:
                    # Adiciona todos os vértices intermediários (exceto o primeiro, já está na lista)
                    caminho_completo.extend(str(v) for v in caminho[1:])
                    break  # achou o caminho, não precisa procurar mais

        caminho_str = " -> ".join(caminho_completo)

        print(f"{alg} ; INICIO: {inicio} ; ATIVOS: {", ".join(str(a) for a in ativos)} ; MELHOR ROTA ENTRE ATIVOS: {" -> ".join(str(a[0]) for a in melhor_rota)} ; CAMINHO TOTAL: {caminho_str}")
        # Exemplo de uso na planilha:
        pagina.append([0, inicio, melhor_rota[-1][1], ", ".join(str(a) for a in ativos), " -> ".join(str(v[0]) for v in melhor_rota), caminho_str, tempo_Rota])
    # Garante que o diretório existe
    planilha.save(nome_arquivo_saida.strip())

def carrega_media_testes():
    global graph_dist, graph_Coordenadas, qtd_vertices, ativos, inicio, distancia_entre_ativos, melhor_rota, caminho_entre_ativos, alg, iteracoes
    
    #algoritmos = ['BFS', 'DFS', 'BCU', 'A_Estrela_Euclidiano', 'A_Estrela_Haversiano']
    planilha = openpyxl.Workbook()
    del planilha['Sheet']
    planilha.create_sheet(alg)

    pagina = planilha[alg]
    pagina.sheet_format.baseColWidth = 30
    pagina.append(['Índice', 'Distancia', 'max', 'min', 'Tempo'])

    sementes = [141592,653589,793238,462643,383279,502884,197169,399375,105820,974944,592307,816406,286208,998628,34825,342117,67982,148086,513282,306647,93844,609550,582231,725359,408128,481117,450284,102701,938521,105559]
    
    for c in range(10): # qtd de testes
        todas_distancias = 0
        todos_tempos = 0

        lista_melores_rotas = []

        qtd_vertices = len(graph_Coordenadas)

        random.seed(sementes[c])
        lista_inicio = []
        lista_ativos = []
        inicio = random.randint(1, qtd_vertices)
        lista_inicio.append(inicio)
        #qtd_ativos_teste = random.randint(1, qtd_vertices-1)
        ativos = random.sample(range(1, qtd_vertices + 1), qtd_ativos)
        lista_ativos.append(ativos)

        for i in sementes:
            random.seed(i)
            melhor_rota = []
            caminho_entre_ativos = []
            distancia_entre_ativos = {}

            # Inicializa as distâncias entre os ativos
            distancia_entre_ativos[inicio] = {}
            for ativo in ativos:
                distancia_entre_ativos[ativo] = {}

            inicio_tempo = time.time()
            melhorar_Rota()
            tempo_Rota = time.time() - inicio_tempo

            print(f"{c} ; {i} ; {ativos} ; {melhor_rota[-1][1]}")

            todas_distancias += melhor_rota[-1][1]
            lista_melores_rotas.append(melhor_rota[-1][1])

            todos_tempos += tempo_Rota
            #print(f"{c} ; {i} = {melhor_rota[-1][1]} ; {tempo_Rota}")

        pagina.append([c, todas_distancias / len(sementes), max(lista_melores_rotas), min(lista_melores_rotas), todos_tempos / len(sementes)])
    texto = nome_arquivo_saida.split("\n")
    planilha.save(f"{texto[0]}")

if __name__ == "__main__":
    qtd_ativos = int(sys.argv[1])
    nome_arquivo_entrada = sys.argv[2]
    nome_arquivo_saida = sys.argv[3]
    alg = sys.argv[4].strip()
    Carrega_Dados()
    #carrega_media_testes()
    carrega_teste()
